# -*- coding: utf-8 -*-
"""Разбор таблицы событий → content/events.json.

Что делает: забирает таблицу (по публичной ссылке Яндекс Диска или из локального
файла), проверяет каждую строку, разворачивает повторяющиеся события по датам,
считает адреса страниц и складывает всё в один json, из которого потом собираются
страницы сайта.

Главное правило: НИ ОДНА плохая строка не должна ронять сборку. Строка с ошибкой
выбрасывается и попадает в отчёт, остальное едет на сайт. Если таблица недоступна
целиком — сборка завершается ошибкой ДО записи файла, и на сайте остаётся прежняя
афиша. Испортить сайт неудачной правкой в таблице нельзя.

Токенов не требует: публичная ссылка Яндекс Диска отдаётся без авторизации.
Проверено 17.08.2026 — это осознанное требование, а не случайность: секретам
в этом репозитории делать нечего.

    python3 tools/events/sobrat.py --file "Афиша Ясны — события.xlsx"
    python3 tools/events/sobrat.py --public-key https://disk.yandex.ru/d/XXXX
"""
import argparse
import json
import re
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta

API = "https://cloud-api.yandex.net/v1/disk/public/resources/download"

UPRAVLENIYA = {
    "Ясна-Школа": "yasna-shkola", "Воспитание и Образование": "vospitanie",
    "Александрия": "alexandria", "Неглинка": "neglinka", "Граника": "granika",
    "Астроневод": "astronevod", "Ясные маршруты": "marshruty",
    "ЛитПроСвет": "litprosvet", "Извод": "izvod", "Джива": "dzhiva",
    "Парад Красоты": "parad", "Herbaldika": "geraldika",
}
TIPY = {"натурный урок", "занятие", "лекция", "встреча", "онлайн-встреча",
        "прогулка", "экспедиция", "праздник", "курс"}
ZAPIS = {"обязательна", "желательна", "не нужна"}
POVTOR_SHAG = {"еженедельно": 7, "раз в две недели": 14}
NA_SAYT = {"опубликовано", "мест нет", "отменено"}

MESYATSY = ["января", "февраля", "марта", "апреля", "мая", "июня",
            "июля", "августа", "сентября", "октября", "ноября", "декабря"]

# Порядок колонок задан таблицей, которую заполняют руководители.
# Если колонку добавят или переставят — разбор идёт по ЗАГОЛОВКУ, а не
# по номеру: иначе одна вставленная колонка сдвинула бы всю афишу.
POLYA = {
    "Состояние": "sostoyanie", "Название": "nazvanie", "Управление": "upravlenie",
    "Тип": "tip", "Дата": "data", "Дата окончания": "data_konca", "Время": "vremya",
    "Город": "gorod", "Точка сбора": "tochka", "Описание": "opisanie",
    "Стоимость": "stoimost", "Запись": "zapis", "Ведёт": "vedet",
    "Для кого": "dlya_kogo", "Что взять": "chto_vzyat", "Мест": "mest",
    "Повтор": "povtor", "Повтор до": "povtor_do",
    "Страница урока на сайте": "stranica", "Комментарий оргкомитету": "kommentariy",
}
OBYAZATELNYE = ["sostoyanie", "nazvanie", "upravlenie", "tip", "data", "vremya",
                "gorod", "tochka", "opisanie", "stoimost", "zapis"]

TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "c",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e",
    "ю": "yu", "я": "ya",
}


def slug(text):
    out = "".join(TRANSLIT.get(ch, ch) for ch in text.lower())
    out = re.sub(r"[^a-z0-9]+", "-", out).strip("-")
    return re.sub(r"-{2,}", "-", out)[:60]


def kletka(v):
    """Ячейка → строка. Даты и числа Excel приводим к тому виду, в котором их
    писал человек: иначе 04.10.2026 превращается в 2026-10-04 00:00:00."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def skachat(public_key, kuda):
    """Публичная ссылка Яндекс Диска → файл на диске. Без токена."""
    url = API + "?" + urllib.parse.urlencode({"public_key": public_key})
    with urllib.request.urlopen(url, timeout=30) as r:
        href = json.load(r)["href"]
    with urllib.request.urlopen(href, timeout=120) as r, open(kuda, "wb") as f:
        f.write(r.read())
    return kuda


def razobrat_datu(s):
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def razobrat_vremya(s):
    m = re.fullmatch(r"(\d{1,2})[:.](\d{2})", s)
    if not m:
        return None
    ch, mn = int(m.group(1)), int(m.group(2))
    return f"{ch:02d}:{mn:02d}" if 0 <= ch < 24 and 0 <= mn < 60 else None


def proverit(row, nomer, segodnya):
    """Строка таблицы → (событие или None, список замечаний)."""
    zam = []
    sost = row["sostoyanie"].lower()

    if sost not in {"черновик", "опубликовано", "мест нет", "отменено"}:
        return None, [f"строка {nomer}: непонятное состояние «{row['sostoyanie']}»"]
    if sost not in NA_SAYT:
        return None, []                      # черновик — молча мимо, это не ошибка

    onlayn = row["tip"] == "онлайн-встреча"
    for p in OBYAZATELNYE:
        if onlayn and p in ("gorod", "tochka"):
            continue
        if not row[p]:
            zam.append(f"строка {nomer}: не заполнено «{p}»")

    if row["upravlenie"] and row["upravlenie"] not in UPRAVLENIYA:
        zam.append(f"строка {nomer}: управления «{row['upravlenie']}» нет в списке")
    if row["tip"] and row["tip"] not in TIPY:
        zam.append(f"строка {nomer}: тип «{row['tip']}» не из списка")
    if row["zapis"] and row["zapis"] not in ZAPIS:
        zam.append(f"строка {nomer}: запись «{row['zapis']}» не из списка")

    d = razobrat_datu(row["data"])
    if row["data"] and not d:
        zam.append(f"строка {nomer}: дата «{row['data']}» не разбирается, нужен вид 04.10.2026")
    t = razobrat_vremya(row["vremya"])
    if row["vremya"] and not t:
        zam.append(f"строка {nomer}: время «{row['vremya']}» не разбирается, нужен вид 12:00")

    # Разметка в тексте: таблица открыта на редактирование двенадцати людям,
    # и всё, что оттуда приходит, попадает в HTML страницы.
    for p in ("nazvanie", "opisanie", "tochka", "dlya_kogo", "chto_vzyat"):
        if re.search(r"<[a-z/!]", row[p], re.I):
            zam.append(f"строка {nomer}: в поле «{p}» разметка, её нельзя")

    if zam:
        return None, zam

    if d < segodnya and sost == "опубликовано":
        return None, [f"строка {nomer}: «{row['nazvanie']}» помечено опубликованным, "
                      f"но {row['data']} уже прошло"]

    ev = dict(row)
    ev["data"] = d
    ev["vremya"] = t
    ev["upr_slug"] = UPRAVLENIYA.get(row["upravlenie"], "")
    ev["konec"] = razobrat_datu(row["data_konca"]) if row["data_konca"] else None
    return ev, []


def razvernut(ev, nomer):
    """Одна строка с повтором → несколько событий по датам."""
    shag = POVTOR_SHAG.get(ev["povtor"], 0)
    if not shag:
        return [ev], []
    do = razobrat_datu(ev["povtor_do"])
    if not do:
        return [ev], [f"строка {nomer}: стоит повтор «{ev['povtor']}», но «Повтор до» "
                      f"не заполнено — событие заведено одной датой"]
    out, d, predel = [], ev["data"], 0
    while d <= do and predel < 200:
        k = dict(ev)
        k["data"] = d
        out.append(k)
        d += timedelta(days=shag)
        predel += 1
    return out, []


def sobrat(put, segodnya):
    from openpyxl import load_workbook
    wb = load_workbook(put, data_only=True)
    if "События" not in wb.sheetnames:
        raise SystemExit("в таблице нет листа «События» — разбирать нечего")
    ws = wb["События"]

    zagolovki = {}
    for c in ws[1]:
        name = kletka(c.value)
        if name in POLYA:
            zagolovki[POLYA[name]] = c.column
    ne_hvataet = [n for n, k in POLYA.items() if k not in zagolovki]
    if ne_hvataet:
        raise SystemExit("в таблице нет колонок: " + ", ".join(ne_hvataet))

    sobytiya, zamechaniya, vsego, chernovikov = [], [], 0, 0
    for r in range(2, ws.max_row + 1):
        row = {k: kletka(ws.cell(row=r, column=col).value) for k, col in zagolovki.items()}
        if not any(row.values()):
            continue
        vsego += 1
        if row["sostoyanie"].lower() == "черновик":
            chernovikov += 1
        ev, zam = proverit(row, r, segodnya)
        zamechaniya += zam
        if not ev:
            continue
        kopii, zam2 = razvernut(ev, r)
        zamechaniya += zam2
        sobytiya += kopii

    # адреса страниц: имя + дата, чтобы у повторов не совпадали
    zanyato = set()
    for ev in sobytiya:
        d = ev["data"]
        baza = f"{slug(ev['nazvanie'])}-{d.day}-{slug(MESYATSY[d.month - 1])}"
        adres, n = baza, 2
        while adres in zanyato:
            adres, n = f"{baza}-{n}", n + 1
        zanyato.add(adres)
        ev["adres"] = adres

    sobytiya.sort(key=lambda e: (e["data"], e["vremya"]))
    return sobytiya, zamechaniya, vsego, chernovikov


def v_json(sobytiya, segodnya):
    out = []
    for e in sobytiya:
        out.append({
            "adres": e["adres"], "sostoyanie": e["sostoyanie"].lower(),
            "nazvanie": e["nazvanie"], "upravlenie": e["upravlenie"],
            "upr_slug": e["upr_slug"], "tip": e["tip"],
            "nachalo": f"{e['data'].isoformat()}T{e['vremya']}:00+03:00",
            "konec": e["konec"].isoformat() if e["konec"] else None,
            "proshlo": e["data"] < segodnya,
            "gorod": e["gorod"], "tochka": e["tochka"], "opisanie": e["opisanie"],
            "stoimost": e["stoimost"], "zapis": e["zapis"], "vedet": e["vedet"],
            "dlya_kogo": e["dlya_kogo"], "chto_vzyat": e["chto_vzyat"],
            "mest": e["mest"], "stranica": e["stranica"],
        })
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file")
    p.add_argument("--public-key")
    p.add_argument("--out", default="content/events.json")
    p.add_argument("--segodnya")
    p.add_argument("--dry-run", action="store_true")
    a = p.parse_args()

    segodnya = razobrat_datu(a.segodnya) if a.segodnya else date.today()
    put = a.file
    if a.public_key:
        put = skachat(a.public_key, "/tmp/afisha.xlsx")

    sobytiya, zamechaniya, vsego, chernovikov = sobrat(put, segodnya)

    print(f"строк в таблице: {vsego}   черновиков: {chernovikov}")
    print(f"событий на сайт: {len(sobytiya)}")
    if zamechaniya:
        print(f"\nОТКЛОНЕНО, {len(zamechaniya)} замечаний:")
        for z in zamechaniya:
            print("  •", z)
    if sobytiya:
        print("\nПОЙДЁТ НА САЙТ:")
        for e in sobytiya:
            print(f"  {e['data'].strftime('%d.%m')} {e['vremya']}  {e['nazvanie'][:44]:<44}"
                  f"  /sobytiya/{e['adres']}/")

    if not a.dry_run:
        with open(a.out, "w", encoding="utf-8") as f:
            json.dump({"sobrano": segodnya.isoformat(),
                       "sobytiya": v_json(sobytiya, segodnya)},
                      f, ensure_ascii=False, indent=1)
        print(f"\nзаписано: {a.out}")
    # Замечания НЕ роняют сборку: отклонённая строка просто не едет на сайт,
    # а всё остальное выкладывается. Ненулевой код здесь остановил бы выкладку
    # из-за одной опечатки в чужой строке. Нечитаемая таблица — другое дело,
    # там SystemExit срабатывает выше, ДО записи файла.
    return 0


if __name__ == "__main__":
    sys.exit(main())
